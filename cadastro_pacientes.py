import openpyxl

continuar_prog = True
pacientes = []

def carregar_dados():
    try:
        workbook = openpyxl.load_workbook('pacientes.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            pacientes.append(list(row))
    except FileNotFoundError:
        print("Arquivo 'pacientes.xlsx' não encontrado. Será criado um novo arquivo.")

def salvar_dados():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Nome', 'Idade', 'Número', 'Endereço'])
    for paciente in pacientes:
        sheet.append(paciente)
    workbook.save('pacientes.xlsx')

def listar_pacientes():
    if not pacientes:
        print("Nenhum paciente cadastrado.")
    else:
        for i, paciente in enumerate(pacientes, 1):
            print(f"{i}. Nome: {paciente[0]}, Idade: {paciente[1]}, Número: {paciente[2]}, Endereço: {paciente[3]}")

def inserir_paciente():
    print('Insira os dados do paciente:')
    nome_paciente = input('Nome: ')
    idade_paciente = input('Idade: ')
    numero_paciente = input('Numero: ')
    endereco_paciente = input('Endereço: ')
    pacientes.append([nome_paciente, idade_paciente, numero_paciente, endereco_paciente])
    print("Paciente cadastrado com sucesso!")

def alterar_paciente():
    listar_pacientes()
    if not pacientes:
        return

    try:
        indice_paciente = int(input("Digite o número do paciente que deseja alterar: ")) - 1
        if 0 <= indice_paciente < len(pacientes):
            print(f"Alterando dados do paciente {pacientes[indice_paciente][0]}:")
            pacientes[indice_paciente][0] = input('Novo nome: ')
            pacientes[indice_paciente][1] = input('Nova idade: ')
            pacientes[indice_paciente][2] = input('Novo número: ')
            pacientes[indice_paciente][3] = input('Novo endereço: ')
            print("Dados do paciente atualizados com sucesso!")
        else:
            print("Índice de paciente inválido.")
    except ValueError:
        print("Valor inválido. Digite um número válido.")

def remover_paciente():
    listar_pacientes()
    if not pacientes:
        return

    try:
        indice_paciente = int(input("Digite o número do paciente que deseja remover: ")) - 1
        if 0 <= indice_paciente < len(pacientes):
            nome_paciente = pacientes[indice_paciente][0]
            del pacientes[indice_paciente]
            print(f"Paciente '{nome_paciente}' removido com sucesso!")
        else:
            print("Índice de paciente inválido.")
    except ValueError:
        print("Valor inválido. Digite um número válido.")

carregar_dados()

while continuar_prog:
    print('\n[1] Inserir paciente')
    print('[2] Alterar paciente')
    print('[3] Remover paciente')
    print('[4] Listar pacientes')
    print('[0] Sair')
    opcao_entrada = input('Selecione uma opção: ')

    try:
        opcao = int(opcao_entrada)
    except ValueError:
        os.system('cls' if os.name == 'nt' else 'clear')
        print('Você digitou uma opção inválida!')
        continue

    if opcao == 1:
        inserir_paciente()
    elif opcao == 2:
        alterar_paciente()
    elif opcao == 3:
        remover_paciente()
    elif opcao == 4:
        listar_pacientes()
    elif opcao == 0:
        salvar_dados()
        print("Encerrando o programa.")
        continuar_prog = False
    else:
        os.system('cls' if os.name == 'nt' else 'clear')
        print('Opção inválida!')
